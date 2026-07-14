import os
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart import BarChart, Reference

def generate_automated_report(csv_file_path, output_excel_path):
    print("Reading and processing data...")
    # 1. Read data using pandas
    df = pd.read_csv(csv_file_path)
    df["Date"] = pd.to_datetime(df["Date"])
    df = df.sort_values("Date").reset_index(drop=True)
    
    # Calculate revenue, cost, and profit
    df["Revenue"] = df["Units_Sold"] * df["Unit_Price"]
    df["Cost"] = df["Units_Sold"] * df["Unit_Cost"]
    df["Profit"] = df["Revenue"] - df["Cost"]
    
    # Aggregate data monthly for the dashboard
    df["Month"] = df["Date"].dt.strftime("%Y-%m")
    monthly_summary = df.groupby("Month")[["Revenue", "Cost", "Profit"]].sum().reset_index()

    # Aggregate data by product categories
    product_summary = df.groupby(["Category", "Product"])[["Units_Sold", "Revenue", "Profit"]].sum().reset_index()
    product_summary["Margin_%"] = product_summary["Profit"] / product_summary["Revenue"]

    # 2. Build the styled Excel workbook using openpyxl
    wb = openpyxl.Workbook()
    
    # Color Palette and Fonts (Professional Navy and Slate Theme)
    navy_header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    zebra_fill = PatternFill(start_color="F2F5F9", end_color="F2F5F9", fill_type="solid")
    kpi_card_fill = PatternFill(start_color="F2F5F9", end_color="F2F5F9", fill_type="solid")
    
    font_title = Font(name="Segoe UI", size=16, bold=True, color="1F4E78")
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_bold = Font(name="Segoe UI", size=11, bold=True)
    font_regular = Font(name="Segoe UI", size=11)
    font_kpi_label = Font(name="Segoe UI", size=9, bold=True, color="595959")
    font_kpi_val = Font(name="Segoe UI", size=18, bold=True, color="1F4E78")
    
    thin_border_side = Side(border_style="thin", color="D9D9D9")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    double_bottom_border = Border(bottom=Side(style='double'), top=Side(style='thin', color="A6A6A6"))

    # ================= TAB 1: Executive Dashboard =================
    ws_dash = wb.active
    ws_dash.title = "Executive Dashboard"
    ws_dash.views.sheetView[0].showGridLines = True  # Ensure gridlines are visible

    # Dashboard Title
    ws_dash["B2"] = "Corporate Performance Dashboard"
    ws_dash["B2"].font = font_title

    # Setup KPI Cards
    kpis = [
        {"range_lbl": "B4:C4", "range_val": "B5:C5", "lbl": "TOTAL REVENUE", "val": f"${df['Revenue'].sum():,.2f}"},
        {"range_lbl": "E4:F4", "range_val": "E5:F5", "lbl": "TOTAL PROFIT", "val": f"${df['Profit'].sum():,.2f}"},
        {"range_lbl": "H4:I4", "range_val": "H5:I5", "lbl": "PROFIT MARGIN %", "val": f"{(df['Profit'].sum() / df['Revenue'].sum())*100:.1f}%"}
    ]
    
    for kpi in kpis:
        ws_dash.merge_cells(kpi["range_lbl"])
        ws_dash.merge_cells(kpi["range_val"])
        lbl_cell = ws_dash[kpi["range_lbl"].split(":")[0]]
        val_cell = ws_dash[kpi["range_val"].split(":")[0]]
        
        lbl_cell.value = kpi["lbl"]
        lbl_cell.font = font_kpi_label
        lbl_cell.alignment = Alignment(horizontal="center")
        
        val_cell.value = kpi["val"]
        val_cell.font = font_kpi_val
        val_cell.alignment = Alignment(horizontal="center")
        
        # Style KPI borders and fills
        cols = kpi["range_lbl"].split(":")
        start_col, end_col = cols[0][0], cols[1][0]
        start_idx = openpyxl.utils.column_index_from_string(start_col)
        end_idx = openpyxl.utils.column_index_from_string(end_col)
        for r in range(4, 6):
            for c in range(start_idx, end_idx + 1):
                cell = ws_dash.cell(row=r, column=c)
                cell.fill = kpi_card_fill
                cell.border = thin_border

    # Monthly Performance Table Headers
    ws_dash["B7"] = "Month"
    ws_dash["C7"] = "Revenue"
    ws_dash["D7"] = "Profit"
    for col in ["B7", "C7", "D7"]:
        ws_dash[col].font = font_header
        ws_dash[col].fill = navy_header_fill
        ws_dash[col].alignment = Alignment(horizontal="center")

    row_idx = 8
    for idx, row in monthly_summary.iterrows():
        ws_dash.cell(row=row_idx, column=2, value=row["Month"]).alignment = Alignment(horizontal="center")
        ws_dash.cell(row=row_idx, column=3, value=row["Revenue"]).number_format = '$#,##0'
        ws_dash.cell(row=row_idx, column=4, value=row["Profit"]).number_format = '$#,##0'
        
        for c in range(2, 5):
            cell = ws_dash.cell(row=row_idx, column=c)
            cell.font = font_regular
            cell.border = thin_border
            if row_idx % 2 == 1:
                cell.fill = zebra_fill
        row_idx += 1

    # Total Row
    ws_dash.cell(row=row_idx, column=2, value="Total").font = font_bold
    ws_dash.cell(row=row_idx, column=2).alignment = Alignment(horizontal="center")
    ws_dash.cell(row=row_idx, column=3, value=f"=SUM(C8:C{row_idx-1})").font = font_bold
    ws_dash.cell(row=row_idx, column=4, value=f"=SUM(D8:D{row_idx-1})").font = font_bold
    ws_dash.cell(row=row_idx, column=3).number_format = '$#,##0'
    ws_dash.cell(row=row_idx, column=4).number_format = '$#,##0'
    for c in range(2, 5):
        ws_dash.cell(row=row_idx, column=c).border = double_bottom_border

    # Add Chart to Dashboard
    chart = BarChart()
    chart.type = "col"
    chart.title = "Monthly Revenue vs Profit"
    chart.y_axis.title = "Amount ($)"
    chart.x_axis.title = "Month"
    chart_data = Reference(ws_dash, min_col=3, min_row=7, max_col=4, max_row=row_idx-1)
    chart_cats = Reference(ws_dash, min_col=2, min_row=8, max_row=row_idx-1)
    chart.add_data(chart_data, titles_from_data=True)
    chart.set_categories(chart_cats)
    chart.height = 10
    chart.width = 15
    ws_dash.add_chart(chart, "F7")

    # ================= TAB 2: Product Metrics =================
    ws_prod = wb.create_sheet(title="Product Metrics")
    ws_prod.views.sheetView[0].showGridLines = True
    ws_prod["A1"] = "Product Line Analysis"
    ws_prod["A1"].font = font_title

    headers_prod = ["Category", "Product", "Units Sold", "Revenue", "Profit", "Margin %"]
    for c_idx, h in enumerate(headers_prod, 1):
        cell = ws_prod.cell(row=3, column=c_idx, value=h)
        cell.font = font_header
        cell.fill = navy_header_fill
        cell.alignment = Alignment(horizontal="center" if c_idx in [1, 2, 6] else "right")

    p_row = 4
    for idx, row in product_summary.iterrows():
        ws_prod.cell(row=p_row, column=1, value=row["Category"]).font = font_bold
        ws_prod.cell(row=p_row, column=2, value=row["Product"]).font = font_regular
        ws_prod.cell(row=p_row, column=3, value=row["Units_Sold"]).number_format = '#,##0'
        ws_prod.cell(row=p_row, column=4, value=row["Revenue"]).number_format = '$#,##0'
        ws_prod.cell(row=p_row, column=5, value=row["Profit"]).number_format = '$#,##0'
        ws_prod.cell(row=p_row, column=6, value=row["Margin_%"]).number_format = '0.0%'
        
        for c in range(1, 7):
            cell = ws_prod.cell(row=p_row, column=c)
            if c != 1:
                cell.font = font_regular
            cell.border = thin_border
            if p_row % 2 == 1:
                cell.fill = zebra_fill
        p_row += 1

    # Product Grand Total
    ws_prod.cell(row=p_row, column=1, value="Grand Total").font = font_bold
    ws_prod.cell(row=p_row, column=3, value=f"=SUM(C4:C{p_row-1})").font = font_bold
    ws_prod.cell(row=p_row, column=4, value=f"=SUM(D4:D{p_row-1})").font = font_bold
    ws_prod.cell(row=p_row, column=5, value=f"=SUM(E4:E{p_row-1})").font = font_bold
    ws_prod.cell(row=p_row, column=6, value=f"=E{p_row}/D{p_row}").font = font_bold
    
    ws_prod.cell(row=p_row, column=3).number_format = '#,##0'
    ws_prod.cell(row=p_row, column=4).number_format = '$#,##0'
    ws_prod.cell(row=p_row, column=5).number_format = '$#,##0'
    ws_prod.cell(row=p_row, column=6).number_format = '0.0%'
    for c in range(1, 7):
        ws_prod.cell(row=p_row, column=c).border = double_bottom_border

    # Conditional Formatting (Soft green highlight for Margin % > 40%)
    green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    green_font = Font(color="375623", bold=True, name="Segoe UI")
    ws_prod.conditional_formatting.add(
        f"F4:F{p_row-1}",
        CellIsRule(operator='greaterThan', formula=['0.4'], stopIfTrue=True, fill=green_fill, font=green_font)
    )

    # ================= TAB 3: Raw Transaction Ledger =================
    ws_raw = wb.create_sheet(title="Raw Transaction Ledger")
    ws_raw.views.sheetView[0].showGridLines = True
    ws_raw.freeze_panes = "A2"  # Freeze top row for easy scrolling

    headers_raw = ["Transaction ID", "Date", "Region", "Category", "Product", "Units Sold", "Unit Price", "Unit Cost", "Revenue", "Cost", "Profit"]
    for c_idx, h in enumerate(headers_raw, 1):
        cell = ws_raw.cell(row=1, column=c_idx, value=h)
        cell.font = font_header
        cell.fill = navy_header_fill
        cell.alignment = Alignment(horizontal="center")

    r_row = 2
    for idx, row in df.iterrows():
        ws_raw.cell(row=r_row, column=1, value=row["Transaction_ID"]).alignment = Alignment(horizontal="center")
        ws_raw.cell(row=r_row, column=2, value=row["Date"].strftime("%Y-%m-%d")).alignment = Alignment(horizontal="center")
        ws_raw.cell(row=r_row, column=3, value=row["Region"]).alignment = Alignment(horizontal="center")
        ws_raw.cell(row=r_row, column=4, value=row["Category"])
        ws_raw.cell(row=r_row, column=5, value=row["Product"])
        ws_raw.cell(row=r_row, column=6, value=row["Units_Sold"]).number_format = '#,##0'
        ws_raw.cell(row=r_row, column=7, value=row["Unit_Price"]).number_format = '$#,##0'
        ws_raw.cell(row=r_row, column=8, value=row["Unit_Cost"]).number_format = '$#,##0'
        ws_raw.cell(row=r_row, column=9, value=row["Revenue"]).number_format = '$#,##0'
        ws_raw.cell(row=r_row, column=10, value=row["Cost"]).number_format = '$#,##0'
        ws_raw.cell(row=r_row, column=11, value=row["Profit"]).number_format = '$#,##0'
        
        for c in range(1, 12):
            cell = ws_raw.cell(row=r_row, column=c)
            cell.font = font_regular
            cell.border = thin_border
            if r_row % 2 == 1:
                cell.fill = zebra_fill
        r_row += 1

    # Auto-adjust column widths based on content length
    for sheet in [ws_dash, ws_prod, ws_raw]:
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if type(cell).__name__ == 'MergedCell':
                    continue
                if cell.value:
                    val_str = str(cell.value)
                    if not val_str.startswith("="):
                        max_len = max(max_len, len(val_str))
            sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # Specific padding for the Dashboard sheet to ensure comfortable layout
    ws_dash.column_dimensions['A'].width = 3
    ws_dash.column_dimensions['B'].width = 15
    ws_dash.column_dimensions['C'].width = 16
    ws_dash.column_dimensions['D'].width = 16

    # Save the workbook
    wb.save(output_excel_path)
    print(f"Success: Report generated at {output_excel_path}")

# Run the script
if __name__ == "__main__":
    raw_data_file = "raw_sales_data.csv"
    output_report_file = "sales_performance_report.xlsx"
    
    if os.path.exists(raw_data_file):
        generate_automated_report(raw_data_file, output_report_file)
    else:
        print(f"Error: Raw data file '{raw_data_file}' not found.")